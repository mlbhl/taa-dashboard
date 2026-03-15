"""
taa_portfolio_optimizer.py의 2단계 배분 로직을 Excel 수식으로 구현하는 스크립트.

Step 1: 자산군 시그널 → 주식/채권 총비중 (SAA 기준, 점수차 × 2.5%p)
Step 2: 개별 시그널 → Base 비례 Tilt로 자산군 내 배분
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── 스타일 정의 ──
ACCENT = "6366F1"
HEADER_FILL = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
HEADER_FONT = Font(name="맑은 고딕", size=10, bold=True, color="475569")
TITLE_FONT = Font(name="맑은 고딕", size=14, bold=True, color="1E293B")
SECTION_FONT = Font(name="맑은 고딕", size=11, bold=True, color=ACCENT)
CELL_FONT = Font(name="맑은 고딕", size=10, color="1E293B")
DIM_FONT = Font(name="맑은 고딕", size=9, color="94A3B8")
PARAM_FONT = Font(name="맑은 고딕", size=10, bold=True, color=ACCENT)
INPUT_FILL = PatternFill(start_color="FFFBEB", end_color="FFFBEB", fill_type="solid")
RESULT_FILL = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
THIN_BORDER = Border(
    bottom=Side(style="thin", color="E2E8F0"),
)
BOTTOM_BORDER = Border(
    bottom=Side(style="medium", color="CBD5E1"),
)

# ── 기본 데이터 ──
DEFAULT_DATA = [
    ("주식", "미국", 49.0, 45.5, "Neutral"),
    ("주식", "유럽", 10.5, 12.6, "Neutral"),
    ("주식", "일본",  3.5,  3.5, "Neutral"),
    ("주식", "중국",  2.1,  3.5, "Neutral"),
    ("주식", "한국",  3.5,  2.1, "Neutral"),
    ("주식", "기타",  1.4,  2.8, "Neutral"),
    ("채권", "미국", 21.0, 18.0, "Neutral"),
    ("채권", "한국",  9.0, 12.0, "Neutral"),
]

TAA_OPTIONS = ["Strong OW", "Overweight", "Neutral", "Underweight", "Strong UW"]
N = len(DEFAULT_DATA)

# 2050 기준 내부비중 (자산군 내 지역별 비율)
EQUITY_INTERNAL = [("주식", "미국", 70), ("주식", "유럽", 15), ("주식", "일본", 5),
                   ("주식", "중국", 3), ("주식", "한국", 5), ("주식", "기타", 2)]
BOND_INTERNAL = [("채권", "미국", 70), ("채권", "한국", 30)]

VINTAGE_CONFIGS = [
    ("2030", 40, 60),
    ("2040", 55, 45),
    ("2060", 90, 10),
]


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_header_row(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BOTTOM_BORDER


def build_sheet(wb):
    """TAA Dashboard 시트 (2단계 배분: 자산군 시그널 + Base 비례 Tilt)"""
    ws = wb.active
    ws.title = "TAA Dashboard"
    set_col_widths(ws, [3, 20, 8, 10, 10, 14, 10, 10, 10, 10, 12, 12, 12])

    # ── Title ──
    ws.merge_cells("B1:H1")
    ws["B1"] = "TAA Dashboard"
    ws["B1"].font = TITLE_FONT

    # ── Parameters ──
    ws["B3"] = "Parameters"
    ws["B3"].font = SECTION_FONT

    params = [
        ("B4", "α (확신도)", "C4", 0.50),
        ("B5", "SAA 가중치 (w)", "C5", 0.50),
        ("B6", "Tilt Rate", "C6", 0.20),
        ("B7", "Tilt Power (p)", "C7", 0.50),
    ]
    for label_cell, label, val_cell, val in params:
        ws[label_cell] = label
        ws[label_cell].font = CELL_FONT
        c = ws[val_cell]
        c.value = val
        c.font = PARAM_FONT
        c.fill = INPUT_FILL
        c.number_format = "0.00"
        c.alignment = Alignment(horizontal="center")

    ws["E4"] = "← 0~1 (보수적 → 적극적)"
    ws["E4"].font = DIM_FONT
    ws["E5"] = "← 0=Peer중심, 1=SAA중심"
    ws["E5"].font = DIM_FONT
    ws["E6"] = "← Base 대비 Tilt 비율"
    ws["E6"].font = DIM_FONT
    ws["E7"] = "← 0=균등, 0.5=감쇠, 1=비례"
    ws["E7"].font = DIM_FONT

    # ── TAA Signal 매핑 참조 (H4:I8 → VLOOKUP 범위) ──
    ws["H3"] = "View Signal 매핑"
    ws["H3"].font = SECTION_FONT
    write_header_row(ws, 4, ["View", "Signal"], start_col=8)
    mapping = [("Strong OW", 2), ("Overweight", 1), ("Neutral", 0), ("Underweight", -1), ("Strong UW", -2)]
    for i, (label, val) in enumerate(mapping):
        ws.cell(row=5 + i, column=8, value=label).font = CELL_FONT
        ws.cell(row=5 + i, column=8).alignment = Alignment(horizontal="center")
        ws.cell(row=5 + i, column=9, value=val).font = PARAM_FONT
        ws.cell(row=5 + i, column=9).alignment = Alignment(horizontal="center")

    # ── 자산군 시그널 (Step 1) ──
    ws["B8"] = "자산군 시그널 (Step 1)"
    ws["B8"].font = SECTION_FONT

    write_header_row(ws, 9, ["자산군", "View", "Score"], start_col=2)

    taa_dv = DataValidation(type="list", formula1='"Strong OW,Overweight,Neutral,Underweight,Strong UW"', allow_blank=True)
    ws.add_data_validation(taa_dv)

    for i, (ac_name, default_taa) in enumerate([("주식", "Neutral"), ("채권", "Neutral")]):
        r = 10 + i
        ws.cell(row=r, column=2, value=ac_name).font = Font(name="맑은 고딕", size=10, bold=True, color=ACCENT)
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=2).border = THIN_BORDER
        c_taa = ws.cell(row=r, column=3, value=default_taa)
        c_taa.font = CELL_FONT
        c_taa.fill = INPUT_FILL
        c_taa.alignment = Alignment(horizontal="center")
        c_taa.border = THIN_BORDER
        taa_dv.add(c_taa)
        # Score = VLOOKUP
        ws.cell(row=r, column=4).value = f"=VLOOKUP(C{r},$H$5:$I$9,2,FALSE)"
        ws.cell(row=r, column=4).font = PARAM_FONT
        ws.cell(row=r, column=4).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=4).border = THIN_BORDER
        ws.cell(row=r, column=4).number_format = "0"

    # Step 1 계산 결과 (오른쪽)
    ws["F8"] = "Step 1 계산"
    ws["F8"].font = SECTION_FONT

    write_header_row(ws, 9, ["항목", "값"], start_col=6)

    # DATA_START는 14이므로, 입력 데이터는 row 15~22
    DATA_START = 14

    # Shift = (주식 Score - 채권 Score) * 2.5
    ws.cell(row=10, column=6, value="Shift(%p)").font = CELL_FONT
    ws.cell(row=10, column=6).alignment = Alignment(horizontal="center")
    ws.cell(row=10, column=6).border = THIN_BORDER
    ws.cell(row=10, column=7).value = "=(D10-D11)*2.5"
    ws.cell(row=10, column=7).font = PARAM_FONT
    ws.cell(row=10, column=7).alignment = Alignment(horizontal="center")
    ws.cell(row=10, column=7).border = THIN_BORDER
    ws.cell(row=10, column=7).number_format = "+0.00;-0.00;0.00"

    # 주식 비중 = MAX(SAA_equity_sum + shift, 1)
    eq_saa_sum = f"SUMIF($B${DATA_START+1}:$B${DATA_START+N},\"주식\",$D${DATA_START+1}:$D${DATA_START+N})"
    bd_saa_sum = f"SUMIF($B${DATA_START+1}:$B${DATA_START+N},\"채권\",$D${DATA_START+1}:$D${DATA_START+N})"

    ws.cell(row=11, column=6, value="주식 비중(%)").font = CELL_FONT
    ws.cell(row=11, column=6).alignment = Alignment(horizontal="center")
    ws.cell(row=11, column=6).border = THIN_BORDER
    # Raw class totals (helper cells)
    ws.cell(row=11, column=8).value = f"=MAX({eq_saa_sum}+$G$10,0)"
    ws.cell(row=11, column=8).font = DIM_FONT
    ws.cell(row=11, column=8).number_format = "0.00"
    ws.cell(row=12, column=8).value = f"=MAX({bd_saa_sum}-$G$10,0)"
    ws.cell(row=12, column=8).font = DIM_FONT
    ws.cell(row=12, column=8).number_format = "0.00"
    # Corrected class totals (subtract excess from larger)
    ws.cell(row=11, column=7).value = "=H11-IF(H11>=H12,MAX(H11+H12-100,0),0)"
    ws.cell(row=11, column=7).font = PARAM_FONT
    ws.cell(row=11, column=7).alignment = Alignment(horizontal="center")
    ws.cell(row=11, column=7).border = THIN_BORDER
    ws.cell(row=11, column=7).number_format = "0.00"

    ws.cell(row=12, column=6, value="채권 비중(%)").font = CELL_FONT
    ws.cell(row=12, column=6).alignment = Alignment(horizontal="center")
    ws.cell(row=12, column=6).border = THIN_BORDER
    ws.cell(row=12, column=7).value = "=H12-IF(H12>H11,MAX(H11+H12-100,0),0)"
    ws.cell(row=12, column=7).font = PARAM_FONT
    ws.cell(row=12, column=7).alignment = Alignment(horizontal="center")
    ws.cell(row=12, column=7).border = THIN_BORDER
    ws.cell(row=12, column=7).number_format = "0.00"

    # ── Region Input Table ──
    ws[f"B{DATA_START - 1}"] = "Region Inputs"
    ws[f"B{DATA_START - 1}"].font = SECTION_FONT

    headers = ["자산", "지역", "SAA(%)", "Peer(%)", "View"]
    write_header_row(ws, DATA_START, headers, start_col=2)

    dv = DataValidation(type="list", formula1='"Strong OW,Overweight,Neutral,Underweight,Strong UW"', allow_blank=True)
    ws.add_data_validation(dv)

    for i, (asset, region, saa, peer, taa) in enumerate(DEFAULT_DATA):
        r = DATA_START + 1 + i
        ws.cell(row=r, column=2, value=asset).font = Font(name="맑은 고딕", size=10, bold=True, color=ACCENT)
        ws.cell(row=r, column=3, value=region).font = Font(name="맑은 고딕", size=10, bold=True)
        c_saa = ws.cell(row=r, column=4, value=saa)
        c_saa.font = CELL_FONT
        c_saa.fill = INPUT_FILL
        c_saa.number_format = "0.0"
        c_saa.alignment = Alignment(horizontal="center")
        c_peer = ws.cell(row=r, column=5, value=peer)
        c_peer.font = CELL_FONT
        c_peer.fill = INPUT_FILL
        c_peer.number_format = "0.0"
        c_peer.alignment = Alignment(horizontal="center")
        c_taa = ws.cell(row=r, column=6, value=taa)
        c_taa.font = CELL_FONT
        c_taa.fill = INPUT_FILL
        c_taa.alignment = Alignment(horizontal="center")
        dv.add(c_taa)
        for col in range(2, 7):
            ws.cell(row=r, column=col).border = THIN_BORDER

    # ── Calculation Table (Step 2) ──
    CALC_START = DATA_START + N + 3
    ws[f"B{CALC_START - 1}"] = "Calculation (Step 2)"
    ws[f"B{CALC_START - 1}"].font = SECTION_FONT

    calc_headers = ["자산", "지역", "SAA", "Peer", "Signal", "Base", "Tilt", "Adj", "Raw", "TAA(%)", "vs Peer", "vs SAA"]
    write_header_row(ws, CALC_START, calc_headers, start_col=2)

    alpha_ref = "$C$4"
    w_ref = "$C$5"
    tilt_rate_ref = "$C$6"
    tilt_power_ref = "$C$7"

    # Raw range refs for per-class SUMIF
    calc_b_range = f"$B${CALC_START+1}:$B${CALC_START+N}"  # 자산 column
    calc_raw_range = f"$J${CALC_START+1}:$J${CALC_START+N}"  # Raw column
    calc_base_range = f"$G${CALC_START+1}:$G${CALC_START+N}"  # Base column (for avg)

    for i in range(N):
        r = CALC_START + 1 + i
        dr = DATA_START + 1 + i

        # 자산, 지역
        ws.cell(row=r, column=2).value = f"=B{dr}"
        ws.cell(row=r, column=2).font = Font(name="맑은 고딕", size=10, bold=True, color=ACCENT)
        ws.cell(row=r, column=3).value = f"=C{dr}"
        ws.cell(row=r, column=3).font = Font(name="맑은 고딕", size=10, bold=True)

        # SAA, Peer
        ws.cell(row=r, column=4).value = f"=D{dr}"
        ws.cell(row=r, column=4).number_format = "0.0"
        ws.cell(row=r, column=5).value = f"=E{dr}"
        ws.cell(row=r, column=5).number_format = "0.0"

        # Signal: VLOOKUP
        taa_cell = f"F{dr}"
        ws.cell(row=r, column=6).value = f"=VLOOKUP({taa_cell},$H$5:$I$9,2,FALSE)"
        ws.cell(row=r, column=6).number_format = "0"

        # Base = w * SAA + (1-w) * Peer
        ws.cell(row=r, column=7).value = f"={w_ref}*D{r}+(1-{w_ref})*E{r}"
        ws.cell(row=r, column=7).number_format = "0.00"

        # Tilt = (Base / avg_Base)^p * avg_Base * tilt_rate
        avg_base_formula = f"SUMIF({calc_b_range},B{r},{calc_base_range})/COUNTIF({calc_b_range},B{r})"
        ws.cell(row=r, column=8).value = f"=(G{r}/({avg_base_formula}))^{tilt_power_ref}*({avg_base_formula})*{tilt_rate_ref}"
        ws.cell(row=r, column=8).number_format = "0.00"

        # Adj = α * Signal * Tilt
        ws.cell(row=r, column=9).value = f"={alpha_ref}*F{r}*H{r}"
        ws.cell(row=r, column=9).number_format = "+0.00;-0.00;0.00"

        # Raw = MAX(Base + Adj, 0)
        ws.cell(row=r, column=10).value = f"=MAX(G{r}+I{r}, 0)"
        ws.cell(row=r, column=10).number_format = "0.00"

        # TAA(%) = Raw / SUMIF(같은 자산군 Raw) * class_total (0이면 0)
        class_ref = f'IF(B{r}="주식",$G$11,$G$12)'
        ws.cell(row=r, column=11).value = (
            f"=IF({class_ref}=0,0,"
            f"J{r}/SUMIF({calc_b_range},B{r},{calc_raw_range})*{class_ref})"
        )
        ws.cell(row=r, column=11).number_format = "0.00"
        ws.cell(row=r, column=11).fill = RESULT_FILL
        ws.cell(row=r, column=11).font = Font(name="맑은 고딕", size=10, bold=True)

        # vs Peer = TAA - Peer
        ws.cell(row=r, column=12).value = f"=K{r}-E{r}"
        ws.cell(row=r, column=12).number_format = "+0.00;-0.00;0.00"

        # vs SAA = TAA - SAA
        ws.cell(row=r, column=13).value = f"=K{r}-D{r}"
        ws.cell(row=r, column=13).number_format = "+0.00;-0.00;0.00"

        for col in range(2, 14):
            c = ws.cell(row=r, column=col)
            if not c.font or c.font == Font():
                c.font = CELL_FONT
            c.alignment = Alignment(horizontal="center")
            c.border = THIN_BORDER

    # ── 합계 행 ──
    sum_row = CALC_START + N + 1
    ws.cell(row=sum_row, column=2, value="합계").font = Font(name="맑은 고딕", size=10, bold=True)
    for col in [4, 5, 7, 10, 11]:
        rng = f"{get_column_letter(col)}{CALC_START+1}:{get_column_letter(col)}{CALC_START+N}"
        ws.cell(row=sum_row, column=col).value = f"=SUM({rng})"
        ws.cell(row=sum_row, column=col).font = Font(name="맑은 고딕", size=10, bold=True)
        ws.cell(row=sum_row, column=col).number_format = "0.0"
        ws.cell(row=sum_row, column=col).alignment = Alignment(horizontal="center")
        ws.cell(row=sum_row, column=col).border = Border(top=Side(style="medium", color="CBD5E1"))

    ws._calc_start = CALC_START

    # ── Range (범위) ──
    RANGE_START = sum_row + 3
    ws[f"B{RANGE_START - 1}"] = "Range (허용 범위)"
    ws[f"B{RANGE_START - 1}"].font = SECTION_FONT

    range_headers = ["자산", "지역", "액티브(%)", "EMP(%)", "TAA(%)", "Half Width", "Low(%)", "High(%)"]
    write_header_row(ws, RANGE_START, range_headers, start_col=2)

    AP_FILL = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")

    for i in range(N):
        r = RANGE_START + 1 + i
        cr = CALC_START + 1 + i
        dr = DATA_START + 1 + i

        ws.cell(row=r, column=2).value = f"=B{cr}"
        ws.cell(row=r, column=2).font = Font(name="맑은 고딕", size=10, bold=True, color=ACCENT)
        ws.cell(row=r, column=3).value = f"=C{cr}"
        ws.cell(row=r, column=3).font = Font(name="맑은 고딕", size=10, bold=True)

        # 액티브 (default=SAA, editable)
        ws.cell(row=r, column=4).value = f"=D{dr}"
        ws.cell(row=r, column=4).number_format = "0.00"
        ws.cell(row=r, column=4).fill = AP_FILL

        # EMP (default=SAA, editable)
        ws.cell(row=r, column=5).value = f"=D{dr}"
        ws.cell(row=r, column=5).number_format = "0.00"
        ws.cell(row=r, column=5).fill = AP_FILL

        # TAA
        ws.cell(row=r, column=6).value = f"=K{cr}"
        ws.cell(row=r, column=6).number_format = "0.00"
        ws.cell(row=r, column=6).fill = RESULT_FILL

        # Half Width
        ws.cell(row=r, column=7).value = f'=IF(F{r}>=20, 7.5, IF(F{r}>=10, 5.0, 2.5))'
        ws.cell(row=r, column=7).number_format = "0.0"

        # Low
        ws.cell(row=r, column=8).value = f"=MAX(F{r}-G{r}, 0)"
        ws.cell(row=r, column=8).number_format = "0.00"
        ws.cell(row=r, column=8).fill = INPUT_FILL

        # High
        ws.cell(row=r, column=9).value = f"=F{r}+G{r}"
        ws.cell(row=r, column=9).number_format = "0.00"
        ws.cell(row=r, column=9).fill = INPUT_FILL

        for col in range(2, 10):
            c = ws.cell(row=r, column=col)
            if not c.font or c.font == Font():
                c.font = CELL_FONT
            c.alignment = Alignment(horizontal="center")
            c.border = THIN_BORDER

    # ── Formula Reference ──
    FORMULA_START = RANGE_START + N + 3
    ws[f"B{FORMULA_START}"] = "Formula Reference"
    ws[f"B{FORMULA_START}"].font = SECTION_FONT
    formulas = [
        "[ Step 1 ] 자산군 비중 결정",
        "  Shift = (주식 Score - 채권 Score) × 2.5%p",
        "  주식 비중 = MAX(SAA_주식합 + Shift, 0),  채권 비중 = MAX(SAA_채권합 - Shift, 0)",
        "  → 합계 보정: 초과분을 큰 쪽에서 차감",
        "",
        "[ Step 2 ] 자산군 내 배분",
        "  1. Signal_i = View → 수치 (SOW=+2, OW=+1, N=0, UW=-1, SUW=-2)",
        "  2. Base_i = w × SAA_i + (1-w) × Peer_i",
        "  3. Tilt_i = (Base_i / avg_Base)^p × avg_Base × Tilt Rate",
        "  4. Adj_i = α × Signal_i × Tilt_i",
        "  5. Raw_i = MAX( Base_i + Adj_i,  0 )",
        "  6. TAA_i = Raw_i / Σ(같은자산군 Raw) × 자산군비중",
        "  7. Range: TAA ≥ 20% → ±7.5%p, ≥ 10% → ±5%p, < 10% → ±2.5%p",
        "",
        "[ 빈티지 전파 ]",
        "  Step 1. 빈티지 자산군비중 = 빈티지SAA ± Shift (동일 적용)",
        "  Step 2. tilt_ratio_i = (Raw_2050_i - SAA_2050_i) / SAA_2050_i",
        "          V_Raw_i = V_SAA_i × (1 + tilt_ratio_i)",
        "          V_TAA = 빈티지자산군비중 × (V_Raw_i / Σ V_Raw_within_class)",
    ]
    for i, f in enumerate(formulas):
        ws.cell(row=FORMULA_START + 1 + i, column=2, value=f).font = Font(name="맑은 고딕", size=9, color="475569")

    # ── Vintage Propagation ──
    # SAA=col D, Raw=col J, calc data starts at CALC_START+1
    vintage_row = FORMULA_START + len(formulas) + 3
    for v_name, v_eq, v_bd in VINTAGE_CONFIGS:
        vintage_row = build_vintage_section(ws, vintage_row, v_name, v_eq, v_bd,
                                            saa_col="D", raw_col="J", calc_start=CALC_START + 1)

    return ws


def build_vintage_section(ws, start_row, vintage_name, equity_pct, bond_pct, saa_col, raw_col, calc_start):
    """빈티지 전파 섹션을 시트에 추가. Step 1(class shift) + Step 2(비례 tilt) 전파."""
    row = start_row
    ws.cell(row=row, column=2, value=f"TDF {vintage_name}  (주식 {equity_pct}% / 채권 {bond_pct}%)")
    ws.cell(row=row, column=2).font = SECTION_FONT
    row += 1

    # Step 1: 빈티지 자산군비중 (class signal shift 반영)
    # G10 = Shift (이미 메인 시트에서 계산됨)
    eq_class_cell = f"H{row}"
    bd_class_cell = f"I{row}"
    ws.cell(row=row, column=2, value="Step 1").font = DIM_FONT
    ws.cell(row=row, column=3, value="V.주식비중").font = DIM_FONT
    ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=4, value="V.채권비중").font = DIM_FONT
    ws.cell(row=row, column=4).alignment = Alignment(horizontal="center")
    # Raw class totals (helper cells J, K)
    raw_eq_cell = f"J{row}"
    raw_bd_cell = f"K{row}"
    ws.cell(row=row, column=10).value = f"=MAX({equity_pct}+$G$10,0)"
    ws.cell(row=row, column=10).font = DIM_FONT
    ws.cell(row=row, column=10).number_format = "0.00"
    ws.cell(row=row, column=11).value = f"=MAX({bond_pct}-$G$10,0)"
    ws.cell(row=row, column=11).font = DIM_FONT
    ws.cell(row=row, column=11).number_format = "0.00"
    # Corrected class totals (subtract excess from larger)
    ws.cell(row=row, column=8).value = f"={raw_eq_cell}-IF({raw_eq_cell}>={raw_bd_cell},MAX({raw_eq_cell}+{raw_bd_cell}-100,0),0)"
    ws.cell(row=row, column=8).font = PARAM_FONT
    ws.cell(row=row, column=8).number_format = "0.00"
    ws.cell(row=row, column=8).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=9).value = f"={raw_bd_cell}-IF({raw_bd_cell}>{raw_eq_cell},MAX({raw_eq_cell}+{raw_bd_cell}-100,0),0)"
    ws.cell(row=row, column=9).font = PARAM_FONT
    ws.cell(row=row, column=9).number_format = "0.00"
    ws.cell(row=row, column=9).alignment = Alignment(horizontal="center")
    step1_row = row
    row += 1

    headers = ["자산", "지역", "V.SAA(%)", "Tilt Ratio", "V.Raw", "TAA(%)", "vs SAA", "Low(%)", "High(%)"]
    write_header_row(ws, row, headers, start_col=2)
    row += 1

    # 빈티지 SAA 계산
    vintage_assets = []
    for asset, region, weight in EQUITY_INTERNAL:
        v_saa = round(equity_pct * weight / 100, 2)
        vintage_assets.append((asset, region, v_saa))
    for asset, region, weight in BOND_INTERNAL:
        v_saa = round(bond_pct * weight / 100, 2)
        vintage_assets.append((asset, region, v_saa))

    data_start_row = row

    # SUMIF ranges for per-class normalization
    v_b_range = f"$B${data_start_row}:$B${data_start_row + N - 1}"
    v_raw_range = f"$F${data_start_row}:$F${data_start_row + N - 1}"

    for i, (asset, region, v_saa) in enumerate(vintage_assets):
        r = row + i
        cr = calc_start + i  # 2050 calc row

        # 자산, 지역
        ws.cell(row=r, column=2, value=asset).font = Font(name="맑은 고딕", size=10, bold=True, color=ACCENT)
        ws.cell(row=r, column=3, value=region).font = Font(name="맑은 고딕", size=10, bold=True)

        # Vintage SAA (정적 값)
        ws.cell(row=r, column=4, value=v_saa)
        ws.cell(row=r, column=4).number_format = "0.00"

        # Tilt Ratio = IF(SAA_2050>=0.5, (Raw_2050-SAA_2050)/SAA_2050, Raw_2050-SAA_2050)
        ws.cell(row=r, column=5).value = f'=IF({saa_col}{cr}>=0.5,({raw_col}{cr}-{saa_col}{cr})/{saa_col}{cr},{raw_col}{cr}-{saa_col}{cr})'
        ws.cell(row=r, column=5).number_format = "+0.00;-0.00;0.00"

        # Vintage Raw = MAX(IF(...), 0)
        ws.cell(row=r, column=6).value = f'=MAX(IF({saa_col}{cr}>=0.5,D{r}*(1+E{r}),D{r}+E{r}),0)'
        ws.cell(row=r, column=6).number_format = "0.00"

        # TAA(%) = per-class normalization (class_total=0이면 0)
        v_class_ref = f'IF(B{r}="주식",{eq_class_cell},{bd_class_cell})'
        ws.cell(row=r, column=7).value = (
            f"=IF({v_class_ref}=0,0,"
            f"F{r}/SUMIF({v_b_range},B{r},{v_raw_range})*{v_class_ref})"
        )
        ws.cell(row=r, column=7).number_format = "0.00"
        ws.cell(row=r, column=7).fill = RESULT_FILL
        ws.cell(row=r, column=7).font = Font(name="맑은 고딕", size=10, bold=True)

        # vs SAA = TAA - V_SAA
        ws.cell(row=r, column=8).value = f"=G{r}-D{r}"
        ws.cell(row=r, column=8).number_format = "+0.00;-0.00;0.00"

        # Half Width → Low, High
        ws.cell(row=r, column=9).value = f"=MAX(G{r}-IF(G{r}>=20,7.5,IF(G{r}>=10,5.0,2.5)),0)"
        ws.cell(row=r, column=9).number_format = "0.00"

        ws.cell(row=r, column=10).value = f"=G{r}+IF(G{r}>=20,7.5,IF(G{r}>=10,5.0,2.5))"
        ws.cell(row=r, column=10).number_format = "0.00"

        for col in range(2, 11):
            c = ws.cell(row=r, column=col)
            if not c.font or c.font == Font():
                c.font = CELL_FONT
            c.alignment = Alignment(horizontal="center")
            c.border = THIN_BORDER

    # 합계 행
    sum_r = data_start_row + N
    ws.cell(row=sum_r, column=2, value="합계").font = Font(name="맑은 고딕", size=10, bold=True)
    for col in [4, 7]:
        rng = f"{get_column_letter(col)}{data_start_row}:{get_column_letter(col)}{data_start_row + N - 1}"
        ws.cell(row=sum_r, column=col).value = f"=SUM({rng})"
        ws.cell(row=sum_r, column=col).font = Font(name="맑은 고딕", size=10, bold=True)
        ws.cell(row=sum_r, column=col).number_format = "0.0"
        ws.cell(row=sum_r, column=col).alignment = Alignment(horizontal="center")
        ws.cell(row=sum_r, column=col).border = Border(top=Side(style="medium", color="CBD5E1"))

    return sum_r + 2  # return next available row


def main():
    wb = openpyxl.Workbook()
    build_sheet(wb)

    output_path = "/home/byoun/projects/taa-dashboard/TAA_Dashboard.xlsx"
    wb.save(output_path)
    print(f"Excel 파일 생성 완료: {output_path}")


if __name__ == "__main__":
    main()
